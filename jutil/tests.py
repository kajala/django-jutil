# pylint: disable=unused-argument,arguments-differ,consider-using-with,unused-variable,unspecified-encoding
import json
import logging
import os
from datetime import datetime, timedelta, date
from decimal import Decimal
from io import BytesIO, StringIO
from os.path import join
from urllib.parse import urlparse
from django.core.management import call_command
from django.utils import timezone
from typing import List
from django.utils.timezone import now
from rest_framework.test import APIClient
from jutil.drf_exceptions import transform_exception_to_drf
from jutil.files import find_file
from jutil.modelfields import SafeCharField, SafeTextField
from jutil.middleware import logger as jutil_middleware_logger, ActivateUserProfileTimezoneMiddleware
import pytz
from django.conf import settings
from django.contrib.admin.models import LogEntry
from django.contrib.auth.models import User
from django.core.exceptions import ValidationError
from rest_framework.exceptions import ValidationError as DRFValidationError, ErrorDetail
from django.core.management.base import CommandParser, BaseCommand  # type: ignore
from django.db import models
from django.http.response import HttpResponse
from django.test import TestCase
from django.test.client import RequestFactory, Client
from django.utils.translation import override, gettext as _, gettext_lazy
from rest_framework.exceptions import NotAuthenticated
from jutil.admin import admin_log, admin_obj_url, admin_obj_link, ModelAdminBase, get_admin_log
from jutil.auth import AuthUserMixin, get_auth_user, get_auth_user_or_none
from jutil.command import get_date_range_by_name, add_date_range_arguments, parse_date_range_arguments, get_command_by_name, get_command_name
from jutil.email import make_email_recipient_list
from jutil.middleware import EnsureOriginMiddleware, LogExceptionMiddleware, EnsureLanguageCookieMiddleware
from jutil.model import (
    is_model_field_changed,
    clone_model,
    get_model_field_label_and_value,
    get_object_or_none,
    wait_object_or_none,
)
from jutil.redis_helpers import redis_set_json, redis_get_json, redis_get_bytes, redis_set_bytes, redis_get_json_or_none, redis_get_bytes_or_none, redis_delete
from jutil.request import get_ip_info
from jutil.responses import CsvResponse
from jutil.testing import TestSetupMixin
from jutil.urls import modify_url
from jutil.xml import xml_to_dict, dict_to_element, _xml_filter_tag_name
from jutil.dates import (
    add_month,
    per_delta,
    per_month,
    this_week,
    next_week,
    this_month,
    last_month,
    last_year,
    last_week,
    yesterday,
    end_of_month,
    this_year,
    get_time_steps,
    TIME_STEP_DAILY,
)
from jutil.format import (
    format_full_name,
    format_xml,
    format_xml_bytes,
    format_timedelta,
    dec1,
    dec2,
    dec3,
    dec4,
    dec5,
    dec6,
    format_table,
    strip_media_root,
    get_media_full_path,
    camel_case_to_underscore,
    underscore_to_camel_case,
    choices_label,
    is_media_full_path,
    capfirst_lazy,
    dec0,
    upper_lazy,
)
from jutil.parse import parse_datetime, parse_bool, parse_datetime_or_none
from jutil.validators import (
    fi_payment_reference_number,
    se_ssn_validator,
    fi_iban_validator,
    se_iban_validator,
    iban_filter_readable,
    email_filter,
    iban_validator,
    iban_bank_info,
    fi_company_org_id_validator,
    email_validator,
    fi_payment_reference_validator,
    iso_payment_reference_validator,
    fi_ssn_age,
    se_clearing_code_bank_info,
    ascii_filter,
    ee_iban_validator,
    be_iban_validator,
    dk_iban_validator,
    dk_iban_bank_info,
    dk_clearing_code_bank_name,
    country_code_sanitizer,
    phone_sanitizer,
    email_sanitizer,
    fi_company_org_id_generator,
    phone_validator,
    passport_filter,
    passport_validator,
    passport_sanitizer,
    country_code_validator,
    validate_country_iban,
    iban_bic,
    validate_country_company_org_id,
    fi_ssn_generator,
    fi_ssn_validator,
    bic_validator,
    iban_generator,
    bic_sanitizer,
    filter_country_company_org_id,
    variable_name_sanitizer,
    is_iban,
    alnum_filter,
)
from xml.etree.ElementTree import Element
from xml.etree import ElementTree as ET  # noqa
from django.contrib import admin


MY_CHOICE_1 = "1"
MY_CHOICE_2 = "2"
MY_CHOICES = (
    (MY_CHOICE_1, "MY_CHOICE_1"),
    (MY_CHOICE_2, "MY_CHOICE_2"),
)

request_factory = RequestFactory()


class DummyLogHandler(logging.Handler):
    msgs: List[str]

    def __init__(self):
        super().__init__()
        self.msgs = []

    def emit(self, record):
        msg = self.format(record)
        self.msgs.append(msg)


def dummy_time_zone_response(obj) -> HttpResponse:
    tz = timezone.get_current_timezone()
    return HttpResponse(str(tz).encode())


def dummy_admin_func_a(modeladmin, request, qs):
    print("dummy_admin_func_a")


def dummy_admin_func_b(modeladmin, request, qs):
    print("dummy_admin_func_b")


def dummy_middleware_get_response(obj) -> HttpResponse:
    return HttpResponse(b"hello content")


class DummyUserProfile:
    timezone = "Europe/Helsinki"


class MyCustomAdmin(ModelAdminBase):
    max_history_length = 5
    actions = (
        dummy_admin_func_b,
        dummy_admin_func_a,
    )

    def get_object(self, request, obj_id):
        return self.model.objects.get(id=obj_id)


class Tests(TestCase, TestSetupMixin):
    def setUp(self):
        self.user = user = self.add_test_user("test@example.com", "test1234")
        assert isinstance(user, User)
        user.is_superuser = True
        user.is_staff = True
        user.save()
        self.client = Client()

    def test_api_client(self):
        api_client = self.create_api_client()
        self.assertTrue(isinstance(api_client, APIClient))

    def test_payment_reference(self):
        self.assertEqual(fi_payment_reference_number("100"), "1009")
        invalids = [
            "89182932u",
            "0000002",
        ]
        for invalid in invalids:
            try:
                fi_payment_reference_number(invalid)
                self.fail("invalid number but passed: {}".format(invalid))
            except ValidationError:
                pass

    def test_format_full_name(self):
        samples = [
            ("Short", "Full Name", "Short Full Name"),
            ("Short Middle Name Is Quite Long", "Full Name", "Short Full Name"),
            ("Short-Middle Name Is Quite Long", "Full Name", "Short Full Name"),
            ("Olga Demi", "Serpuhovitinova-Miettinen", "Olga Serpuhovitinova"),
            ("Olga-Anne Demi", "Serpuhovitinovatsko", "Olga S"),
        ]
        for v in samples:
            limited = format_full_name(v[0], v[1], 20)
            # print('{} {} -> {} (was: {})'.format(v[0], v[1], v[2], limited))
            self.assertEqual(v[2], limited)
        try:
            long_name = "19280309812083091829038190823081208301280381092830182038018203810283021"
            format_full_name(long_name, long_name)
            self.fail("format_full_name failed with long name")
        except Exception:
            pass

    def test_parse_datetime(self):
        pairs = [
            ("2016-06-12", "2016-06-12T00:00:00+00:00"),
            ("2016-06-12T01:00:00", "2016-06-12T01:00:00+00:00"),
            ("2016-06-12 01:00:00", "2016-06-12T01:00:00+00:00"),
            ("2016-06-12T01:00:00+00:00", "2016-06-12T01:00:00+00:00"),
            ("2016-06-12 01:00:00+00:00", "2016-06-12T01:00:00+00:00"),
        ]
        for t_str_input, t_ref_str in pairs:
            # parse_datetime
            t = parse_datetime(t_str_input)
            self.assertTrue(isinstance(t, datetime))
            assert isinstance(t, datetime)
            self.assertEqual(t.isoformat(), t_ref_str)
            # parse_datetime_or_none
            t = parse_datetime_or_none(t_str_input)
            self.assertTrue(isinstance(t, datetime))
            assert isinstance(t, datetime)
            self.assertEqual(t.isoformat(), t_ref_str)

        invalids = [
            "12312313ew",
            "2016-13-05",
        ]
        for t_str_input in invalids:
            # parse_datetime
            try:
                parse_datetime(t_str_input)
                self.fail("{} is not valid input for parse_datetime()".format(t_str_input))
            except Exception:
                pass
            # parse_datetime_or_none
            t = parse_datetime_or_none(t_str_input)
            assert t is None

    def test_add_month(self):
        t = parse_datetime("2016-06-12T01:00:00")
        self.assertTrue(isinstance(t, datetime))
        assert isinstance(t, datetime)
        self.assertEqual(t.isoformat(), "2016-06-12T01:00:00+00:00")
        t2 = add_month(t)
        self.assertEqual(t2.isoformat(), "2016-07-12T01:00:00+00:00")
        t3 = add_month(t, 7)
        self.assertEqual(t3.isoformat(), "2017-01-12T01:00:00+00:00")
        t4 = add_month(t, -1)
        self.assertEqual(t4.isoformat(), "2016-05-12T01:00:00+00:00")
        time_now = datetime(2020, 6, 30, 15, 47, 23, 818646)
        self.assertEqual(add_month(time_now, -4).isoformat(), "2020-02-29T15:47:23.818646")
        self.assertEqual(add_month(time_now, 8).isoformat(), "2021-02-28T15:47:23.818646")
        self.assertEqual(add_month(time_now, 0).isoformat(), "2020-06-30T15:47:23.818646")

    def test_se_ssn(self):
        se_ssn_validator("811228-9874")
        se_ssn_validator("670919-9530")
        with self.assertRaises(ValidationError):
            se_ssn_validator("811228-9873")

    def test_phone_numbers(self):
        try:
            phone_validator("+358456343767")
        except Exception:
            self.fail('phone_validator("+358456343767") should not raise Exception')
        with self.assertRaisesMessage(ValidationError, _("Invalid phone number")):
            phone_validator("214")
        self.assertEqual(phone_sanitizer("214"), "")

    def test_passport(self):
        self.assertEqual(passport_filter("?ADsd-12312dsds"), "ADSD-12312DSDS")
        passport_validator("21412312312")
        with self.assertRaisesMessage(ValidationError, _("Invalid passport number")):
            passport_validator("214")
        self.assertEqual(passport_sanitizer("214"), "")
        self.assertEqual(passport_sanitizer("21412312312"), "21412312312")

    def test_country_code(self):
        for cc in ["FI", "DK", "ES", "SE", "VN"]:
            country_code_validator(cc)
        with self.assertRaisesMessage(ValidationError, _("Invalid country code")):
            country_code_validator("Finland")

    def test_bic(self):
        bic = iban_bic("FI21 1234 5600 0007 85")
        self.assertEqual(bic, "NDEAFIHH")
        self.assertEqual(bic_sanitizer("NDEAFIHH"), "NDEAFIHH")
        self.assertEqual(bic_sanitizer("NDEAFIH"), "")

    def test_org_id(self):
        try:
            validate_country_company_org_id("FI", "2084069-9")
        except Exception:
            self.fail("2084069-9 is valid org id")
        with self.assertRaisesMessage(ValidationError, _("Invalid company organization ID")):
            validate_country_company_org_id("FI", "2084069-8")

    def test_validate_country_iban(self):
        with self.assertRaisesMessage(ValidationError, _("Invalid IBAN account number")):
            validate_country_iban("FI15616515616156", "SE")

    def test_fi_ssn_generator(self):
        self.assertEqual(len(fi_ssn_generator()), 6 + 1 + 4)
        for min_year, max_year in [(1800, 1900), (1900, 2000), (2000, 2050)]:
            for n in range(10):
                ssn = fi_ssn_generator(min_year, max_year)
                try:
                    fi_ssn_age(ssn)
                    fi_ssn_validator(ssn)
                except Exception:
                    self.fail("{} is valid SSN".format(ssn))
        fi_ssn_validator("110305+283X")
        for ssn in ["9999-123F", "271138-670X", "090228+256X"]:
            with self.assertRaisesMessage(ValidationError, _("Invalid personal identification number")):
                fi_ssn_validator(ssn)
        with self.assertRaises(ValidationError):
            fi_ssn_generator(1700, 1799)
        with self.assertRaises(ValidationError):
            fi_ssn_generator(2100, 2200)

    def test_iban(self):
        with self.assertRaisesMessage(ValidationError, _("Invalid IBAN account number")):
            iban_validator("")
        with self.assertRaisesMessage(ValidationError, _("Invalid country code")):
            iban_validator("XX")
        with self.assertRaisesMessage(ValidationError, _("Invalid IBAN account number")):
            iban_validator("FI2112345600000786")
        bic_validator("HELSFIHH")
        bic_validator("HELSFIHHXXX")
        with self.assertRaisesMessage(ValidationError, _("Invalid bank BIC/SWIFT code")):
            bic_validator("HELSFIH")
        with self.assertRaisesMessage(ValidationError, _("Invalid bank BIC/SWIFT code")):
            bic_validator("")
        with self.assertRaisesMessage(ValidationError, _("Invalid bank BIC/SWIFT code")):
            bic_validator("XX123123123112")
        iban_validator("FI2112345600000785")
        iban_validator("SE4550000000058398257466")
        fi_iban_validator("FI2112345600000785")
        se_iban_validator("SE4550000000058398257466")
        ee_iban_validator("EE38 2200 2210 2014 5685")
        be_iban_validator("BE68 5390 0754 7034")
        with self.assertRaises(ValidationError):
            fi_iban_validator("FI2112345600000784")
        with self.assertRaises(ValidationError):
            se_iban_validator("SE4550000000058398257465")
        iban = "FI8847304720017517"
        self.assertEqual(iban_filter_readable(iban), "FI88 4730 4720 0175 17")
        self.assertEqual(iban_filter_readable(""), "")

    def test_urls(self):
        self.assertEqual(
            modify_url("http://yle.fi/uutiset/3-8045550?a=123&b=456", {"a": "123", "b": "456"}),
            "http://yle.fi/uutiset/3-8045550?a=123&b=456",
        )

    def test_email_filter_and_validation(self):
        emails = [
            (" Asdsa@a-a.com ", "asdsa@a-a.com", True),
            ("1asdsa@a-a2.com", "1asdsa@a-a2.com", True),
            (" Asdsa@a-a ", "asdsa@a-a", False),
            (" @a-a2.com", "@a-a2.com", False),
            (" a-a2.com", "a-a2.com", False),
            ("ää-a2@ää-a2.com", "ää-a2@ää-a2.com", False),
            ("aaa.bbbbb@ccc-ddd.fi", "aaa.bbbbb@ccc-ddd.fi", True),
        ]
        for i, o, is_valid in emails:
            # print('email_filter({}) -> {}'.format(i, email_filter(i)))
            self.assertEqual(email_filter(i), o)
            if is_valid:
                email_validator(o)
            else:
                fail = False
                try:
                    email_validator(o)
                except ValidationError:
                    fail = True
                self.assertTrue(fail, "{} is not valid email but passed validation".format(o))

    def test_ip_info(self):
        ip, cc, host = get_ip_info("213.214.146.142")
        self.assertEqual(ip, "213.214.146.142")
        if cc:
            self.assertEqual(cc, "FI")
        if host:
            self.assertEqual(host, "213214146142.edelkey.net")

    def test_parse_xml(self):
        # finvoice_201_example1.xml
        xml_bytes = open(join(settings.BASE_DIR, "data/fi/finvoice_201_example1.xml"), "rb").read()
        data = xml_to_dict(xml_bytes, value_key="value", attribute_prefix="_")
        # pprint(data)
        self.assertEqual(data["_Version"], "2.01")
        self.assertEqual(data["InvoiceRow"][0]["ArticleIdentifier"], "12345")
        self.assertEqual(data["InvoiceRow"][0]["DeliveredQuantity"]["value"], "2")
        self.assertEqual(data["InvoiceRow"][0]["DeliveredQuantity"]["_QuantityUnitCode"], "kpl")
        self.assertEqual(data["InvoiceRow"][1]["ArticleIdentifier"], "123456")

        # parse_xml1.xml
        xml_str = open(join(settings.BASE_DIR, "data/parse_xml1.xml"), "rt").read()
        data = xml_to_dict(xml_str.encode())
        # pprint(data)
        ref_data = {
            "@version": "1.2",
            "A": [
                {"@class": "x", "B": {"@": "hello", "@class": "x2"}},
                {"@class": "y", "B": {"@": "world", "@class": "y2"}},
            ],
            "C": "value node",
        }
        self.assertEqual(ref_data, data)

        # parse_xml1.xml / no attributes
        xml_str = open(join(settings.BASE_DIR, "data/parse_xml1.xml"), "rt").read()
        data = xml_to_dict(xml_str.encode(), parse_attributes=False)
        # pprint(data)
        ref_data = {"A": [{"B": "hello"}, {"B": "world"}], "C": "value node"}
        self.assertEqual(ref_data, data)

        # parse_xml2.xml / no attributes
        xml_str = open(join(settings.BASE_DIR, "data/parse_xml2.xml"), "rt").read()
        data = xml_to_dict(xml_str.encode(), ["VastausLoki", "LuottoTietoMerkinnat"], parse_attributes=False)
        # pprint(data)
        ref_data = {
            "VastausLoki": {
                "KysyttyHenkiloTunnus": "020685-1234",
                "PaluuKoodi": "Palveluvastaus onnistui",
                "SyyKoodi": "1",
            }
        }
        self.assertEqual(ref_data, data)

    def test_dict_to_xml(self):
        data = {
            "Doc": {
                "@version": "1.2",
                "A": [
                    {"@class": "x", "B": {"@": "hello", "@class": "x2"}},
                    {"@class": "y", "B": {"@": "world", "@class": "y2"}},
                ],
                "C": "value node",
                "D": 123,
                "E": ["abc"],
            }
        }
        el = dict_to_element(data)
        assert isinstance(el, Element)
        xml_str = ET.tostring(el, encoding="utf8", method="xml").decode()
        # print(xml_str)  # <Doc version="1.2"><C>value node</C><A class="x"><B class="x2">hello</B></A><A class="y"><B class="y2">world</B></A></Doc>
        data2 = xml_to_dict(xml_str.encode(), document_tag=True, array_tags=["E"], int_tags=["D"])
        # print('')
        # pprint(data)
        # pprint(data2)
        self.assertEqual(data2, data)

    def test_dict_to_xml2(self):
        self.assertEqual(_xml_filter_tag_name("TagName[0]"), "TagName")
        self.assertEqual(_xml_filter_tag_name("TagName[1]"), "TagName")
        self.assertEqual(_xml_filter_tag_name("TagName"), "TagName")
        data = {
            "Doc": {
                "@version": "1.2",
                "A": [
                    {"@class": "x", "B": {"@": "hello", "@class": "x2"}},
                    {"@class": "y", "B": {"@": "world", "@class": "y2"}},
                ],
                "C": "value node",
                "D": 123,
                "E": ["abc"],
                "F": ["line 1", "line 2"],
            }
        }
        el = dict_to_element(data)
        assert isinstance(el, Element)
        xml_str = ET.tostring(el, encoding="utf8", method="xml").decode()
        data2 = xml_to_dict(xml_str.encode(), document_tag=True, array_tags=["E", "F"], int_tags=["D"])
        self.assertEqual(data2, data)

    def test_xml_to_dict(self):
        xml_str = """<?xml version="1.0" encoding="utf-8"?>
<Document>
  <TxsSummry>
    <TtlNtries>
      <NbOfNtries>12</NbOfNtries>
    </TtlNtries>
    <TtlCdtNtries>
      <NbOfNtries>34</NbOfNtries>
      <Sum>1234.56</Sum>
    </TtlCdtNtries>
    <TtlDbtNtries>
      <NbOfNtries>0</NbOfNtries>
      <Sum>0</Sum>
    </TtlDbtNtries>
  </TxsSummry>
</Document>"""
        data = xml_to_dict(xml_str.encode(), document_tag=True, array_tags=[], int_tags=["NbOfNtries"])
        # print('')
        # pprint(data)
        self.assertEqual(data["Document"]["TxsSummry"]["TtlNtries"]["NbOfNtries"], 12)
        self.assertEqual(data["Document"]["TxsSummry"]["TtlCdtNtries"]["NbOfNtries"], 34)

    def test_per_delta(self):
        begin = datetime(2017, 9, 17, 11, 42)
        end = begin + timedelta(days=4)
        ref = [
            (datetime(2017, 9, 17, 11, 42), datetime(2017, 9, 18, 11, 42)),
            (datetime(2017, 9, 18, 11, 42), datetime(2017, 9, 19, 11, 42)),
            (datetime(2017, 9, 19, 11, 42), datetime(2017, 9, 20, 11, 42)),
            (datetime(2017, 9, 20, 11, 42), datetime(2017, 9, 21, 11, 42)),
        ]
        res = per_delta(begin, end, timedelta(days=1))
        self.assertEqual(list(res), ref)

    def test_per_month(self):
        begin = datetime(2017, 9, 1, 0, 0)
        res = list(per_month(begin, begin + timedelta(days=32)))
        ref = [
            (datetime(2017, 9, 1, 0, 0), datetime(2017, 10, 1, 0, 0)),
            (datetime(2017, 10, 1, 0, 0), datetime(2017, 11, 1, 0, 0)),
        ]
        self.assertEqual(list(res), ref)

    def test_dates(self):
        t = datetime(2018, 1, 30)
        b, e = this_week(t)
        self.assertEqual(b, pytz.utc.localize(datetime(2018, 1, 29)))
        self.assertEqual(e, pytz.utc.localize(datetime(2018, 2, 5)))
        b, e = this_month(t)
        self.assertEqual(b, pytz.utc.localize(datetime(2018, 1, 1)))
        self.assertEqual(e, pytz.utc.localize(datetime(2018, 2, 1)))
        b, e = next_week(t)
        self.assertEqual(b, pytz.utc.localize(datetime(2018, 2, 5)))
        self.assertEqual(e, pytz.utc.localize(datetime(2018, 2, 12)))

    def test_named_date_ranges(self):
        t = datetime(2018, 5, 31)
        t_tz = pytz.utc.localize(t)
        named_ranges = [
            ("last_year", last_year(t)),
            ("last_month", last_month(t)),
            ("last_week", last_week(t)),
            ("this_year", this_year(t)),
            ("this_month", this_month(t)),
            ("this_week", this_week(t)),
            ("yesterday", yesterday(t)),
            ("today", yesterday(t + timedelta(hours=24))),
        ]
        day_ranges = [7, 15, 30, 60, 90]
        for days in day_ranges:
            named_ranges.append(("plus_minus_{}d".format(days), (t_tz - timedelta(days=days), t_tz + timedelta(days=days))))
            named_ranges.append(("prev_{}d".format(days), (t_tz - timedelta(days=days), t_tz)))
            named_ranges.append(("next_{}d".format(days), (t_tz, t_tz + timedelta(days=days))))
        for name, res in named_ranges:
            # print('testing', name)
            self.assertEqual(get_date_range_by_name(name, t), res)

    def test_time_steps(self):
        t = parse_datetime("2020-08-24 23:28:36.503174")
        this_week_daily = [
            "2020-08-24T00:00:00+00:00 2020-08-25T00:00:00+00:00",
            "2020-08-25T00:00:00+00:00 2020-08-26T00:00:00+00:00",
            "2020-08-26T00:00:00+00:00 2020-08-27T00:00:00+00:00",
            "2020-08-27T00:00:00+00:00 2020-08-28T00:00:00+00:00",
            "2020-08-28T00:00:00+00:00 2020-08-29T00:00:00+00:00",
            "2020-08-29T00:00:00+00:00 2020-08-30T00:00:00+00:00",
            "2020-08-30T00:00:00+00:00 2020-08-31T00:00:00+00:00",
        ]
        ix = 0
        for begin, end in get_time_steps(TIME_STEP_DAILY, *this_week(t)):
            self.assertEqual("{} {}".format(begin.isoformat(), end.isoformat()), this_week_daily[ix])
            ix += 1

    def test_bank_info(self):
        ac = "FI8847304720017517"
        inf = iban_bank_info(ac)
        self.assertEqual(inf[0], "POPFFI22")
        self.assertEqual(inf[1], "POP-Pankki")

        ac = ""
        inf = iban_bank_info(ac)
        self.assertEqual(inf[0], "")
        self.assertEqual(inf[1], "")

        ac = "BE75270187592710"
        inf = iban_bank_info(ac)
        self.assertEqual(inf[0], "GEBABEBB")
        self.assertEqual(inf[1], "BNP Paribas Fortis")
        ac = "BE58465045170210"
        inf = iban_bank_info(ac)
        self.assertEqual(inf[0], "KREDBEBB")
        self.assertEqual(inf[1], "KBC Bank")
        ac = "BE11000123456748"
        inf = iban_bank_info(ac)
        self.assertEqual(inf[0], "BPOTBEB1")
        self.assertEqual(inf[1], "bpost bank")

    def test_org_id_fi(self):
        valids = [
            "FI01098230",
            "FI-01098230",
            "0109823-0",
            "2084069-9",
        ]
        invalids = [
            "2084069-1",
            "SE2084069-1",
        ]
        co_filtered = [
            ("FI", "FI01098230", "0109823-0"),
            ("FI", "FI-01098230", "0109823-0"),
            ("FI", "0109823-0", "0109823-0"),
            ("FI", "2084069-9", "2084069-9"),
            ("SE", "01098230", "01098230"),
            ("SE", "20840699", "20840699"),
        ]
        for cc, org, val in co_filtered:
            self.assertEqual(filter_country_company_org_id(cc, org), val)
            validate_country_company_org_id(cc, org)
        for valid in valids:
            fi_company_org_id_validator(valid)
        for invalid in invalids:
            try:
                fi_company_org_id_validator(invalid)
                self.fail("{} passed as valid FI-org".format(invalid))
            except ValidationError:
                # print('ok')
                pass
        for n in range(10):
            v0 = fi_company_org_id_generator()
            # print(v0)
            fi_company_org_id_validator(v0)

    def test_reference_number_validators(self):
        valid_fi_refs = [
            "302300",
            "202196",
            "302290",
        ]
        for ref_no in valid_fi_refs:
            fi_payment_reference_validator(ref_no)

        invalid_fi_refs = [
            "302301",
            "202195",
            "302291",
        ]
        for ref_no in invalid_fi_refs:
            try:
                fi_payment_reference_validator(ref_no)
                self.fail("{} should have failed validation".format(ref_no))
            except ValidationError:
                pass

        valid_iso_refs = [
            "RF92 1229",
            "RF11 1232",
            "RF48 1245",
        ]
        for ref_no in valid_iso_refs:
            iso_payment_reference_validator(ref_no)

        invalid_iso_refs = [
            "RF92 1229}",
        ]
        for ref_no in invalid_iso_refs:
            with self.assertRaisesMessage(ValidationError, "Invalid payment reference"):
                iso_payment_reference_validator(ref_no)

    def test_fi_ssn_age(self):
        samples = [
            (date(2018, 12, 20), "231298-965X", 19),
            (date(2018, 12, 22), "231298-965X", 19),
            (date(2018, 12, 23), "231298-965X", 20),
            (date(2018, 12, 24), "231298-965X", 20),
        ]
        for date_now, ssn, age in samples:
            self.assertEqual(
                fi_ssn_age(ssn, date_now),
                age,
                msg="{} age is {} on {} but fi_ssn_age result was {}".format(ssn, age, date_now, fi_ssn_age(ssn, date_now)),
            )

    def test_se_banks(self):
        self.assertEqual(iban_bank_info("SE7912001200012350223035")[1], "Danske Bank")
        self.assertEqual(se_clearing_code_bank_info("6789"), ("Handelsbanken", 9))
        se_iban_validator("SE45 5000 0000 0583 9825 7466")
        with self.assertRaisesMessage(ValidationError, _("Invalid IBAN account number")):
            se_iban_validator("")
        with self.assertRaisesMessage(ValidationError, _("Invalid IBAN account number")):
            se_iban_validator("XX45 5000 0000 0583 9825 7466")
        with self.assertRaisesMessage(ValidationError, _("Invalid IBAN account number")):
            se_iban_validator("SE45 5000 0000 0583 9825")
        self.assertEqual(se_clearing_code_bank_info("9500"), ("Nordea AB", 10))
        an = "957033025420"
        bank_name, acc_digits = se_clearing_code_bank_info(an)
        self.assertEqual(bank_name, "Sparbanken Syd")
        self.assertGreaterEqual(len(an) - 4, acc_digits)

    def test_dk_banks(self):
        an = "DK50 0040 0440 1162 43"
        dk_iban_validator(an)
        bic, name = dk_iban_bank_info(an)
        self.assertEqual(name, "Nordea")
        an = "8114 0008874093"
        name = dk_clearing_code_bank_name(an)
        self.assertEqual(name, "Nykredit Bank")
        an = "DK2520006893703029"
        name = dk_clearing_code_bank_name(an)
        self.assertEqual(name, "Nordea")

    def test_ascii_filter(self):
        pairs = [
            ("Åke va Källe o Öring", "Ake va Kalle o Oring"),
            ("Tôi đang đi mua sắm", "Toi ang i mua sam"),
            ("HELÉN FRANZÉN", "HELEN FRANZEN"),
        ]
        for a, b in pairs:
            self.assertEqual(ascii_filter(a), b, 'ascii_filter("{}") != "{}"'.format(b, ascii_filter(a)))

    def test_l10n(self):
        with override("fi"):
            msg = _("“%(value)s” value has an invalid format. It must be in YYYY-MM-DD HH:MM[:ss[.uuuuuu]][TZ] format.")
            if "muoto ei kelpaa" not in msg:
                print(msg)
            self.assertTrue("muoto ei kelpaa" in msg)

            try:
                parse_bool("hello")
            except ValidationError as e:
                self.assertEqual(str(e), "['hello ei ole yksi valittavissa olevista arvoista']")

    def test_sanitizers(self):
        self.assertEqual(country_code_sanitizer("kods"), "")
        self.assertEqual(country_code_sanitizer("fi"), "FI")
        self.assertEqual(phone_sanitizer("+13146094459"), "+13146094459")
        self.assertEqual(phone_sanitizer("13146094459"), "13146094459")
        self.assertEqual(phone_sanitizer("+13146094459A"), "+13146094459")
        self.assertEqual(email_sanitizer("test@example.com"), "test@example.com")
        self.assertEqual(email_sanitizer("testexample.com"), "")

    def test_format_xml(self):
        assert settings.XMLLINT_PATH, 'add e.g. XMLLINT_PATH = "/usr/bin/xmllint" to settings.py'
        src = "<ApplicationRequest> <CustomerId>1</CustomerId>  <Command>DownloadFileList</Command><Timestamp>2019-11-27T04:32:18.613452+02:00</Timestamp><Environment>PRODUCTION</Environment></ApplicationRequest>"  # noqa
        dst_ref = '<?xml version="1.0"?>\n<ApplicationRequest>\n  <CustomerId>1</CustomerId>\n  <Command>DownloadFileList</Command>\n  <Timestamp>2019-11-27T04:32:18.613452+02:00</Timestamp>\n  <Environment>PRODUCTION</Environment>\n</ApplicationRequest>\n'  # noqa
        dst = format_xml(src)
        self.assertEqual(dst, dst_ref)
        dst = format_xml_bytes(src.encode())
        self.assertEqual(dst, dst_ref.encode())

    def test_parse_sftp(self):
        test_cases = [
            ("sftp://jani@kajala.com", ["jani", None, "kajala.com", ""]),
            ("sftp://jani.kajala:1231!@kajala.com", ["jani.kajala", "1231!", "kajala.com", ""]),
            ("sftp://jani:1231!@kajala.com/my/dir", ["jani", "1231!", "kajala.com", "/my/dir"]),
            ("sftp://jani.kajala:1231!@kajala.com/my/dir", ["jani.kajala", "1231!", "kajala.com", "/my/dir"]),
        ]
        for connection, ref_res in test_cases:
            res = urlparse(connection)
            res_list = [res.username, res.password, res.hostname, res.path]
            self.assertListEqual(res_list, ref_res, 'SFTP connection string "{}" parsed incorrectly'.format(connection))

    def test_admin(self):
        obj = self.user
        admin_log([obj], "Hello, world")
        admin_log([obj], "Hello, world", user=self.user, ip="127.0.0.1")
        admin_log(obj, "Hello, world", user=self.user, ip="127.0.0.1")
        self.assertGreaterEqual(get_admin_log(obj).filter(change_message__contains="Hello, world").count(), 3)
        e = LogEntry.objects.all().filter(object_id=obj.id).last()
        self.assertIsNotNone(e)
        assert isinstance(e, LogEntry)
        self.assertEqual(e.change_message, "Hello, world")
        self.assertEqual(admin_obj_url(obj, "admin:auth_user_change"), "/admin/auth/user/{}/change/".format(obj.id))
        self.assertEqual(admin_obj_url(None, "admin:auth_user_change"), "")
        self.assertEqual(admin_obj_link(None, "admin:auth_user_change"), "")
        self.assertEqual(admin_obj_url(obj), "/admin/auth/user/{}/change/".format(obj.id))
        self.assertEqual(admin_obj_url(e), "/admin/admin/logentry/{}/change/".format(e.id))
        link = admin_obj_link(obj, "User", "admin:auth_user_change")
        self.assertEqual(link, "<a href='/admin/auth/user/{}/change/'>User</a>".format(obj.id))

    def test_cmd_parser(self):
        parser = CommandParser()
        add_date_range_arguments(parser)
        argv = parser.parse_args(["--begin", "2019-06-25", "--end", "2020-02-01"])
        options = argv.__dict__
        begin, end, steps = parse_date_range_arguments(options)
        self.assertEqual(begin, pytz.utc.localize(datetime(2019, 6, 25)))
        self.assertEqual(end, pytz.utc.localize(datetime(2020, 2, 1)))

    def test_format_timedelta(self):
        self.assertEqual(format_timedelta(timedelta(seconds=90)), "1min30s")
        self.assertEqual(format_timedelta(timedelta(seconds=3600 + 90)), "1h1min30s")
        self.assertEqual(format_timedelta(timedelta(seconds=90), minutes_label="min ", seconds_label="s "), "1min 30s")
        self.assertEqual(
            format_timedelta(timedelta(seconds=3600 + 90), hours_label="h ", minutes_label="min ", seconds_label="s "),
            "1h 1min 30s",
        )
        self.assertEqual(format_timedelta(timedelta(seconds=90), seconds_label=""), "1min")
        self.assertEqual(format_timedelta(timedelta(seconds=0.090), seconds_label="s"), "0.090s")

    def test_dec0123456(self):
        self.assertEqual(dec0(Decimal("1.2345")), Decimal("1"))
        self.assertEqual(dec1(Decimal("1.2345678")), Decimal("1.2"))
        self.assertEqual(dec2(Decimal("1.2345678")), Decimal("1.23"))
        self.assertEqual(dec3(Decimal("1.2345678")), Decimal("1.235"))
        self.assertEqual(dec4(Decimal("1.2345678")), Decimal("1.2346"))
        self.assertEqual(dec5(Decimal("1.2345678")), Decimal("1.23457"))
        self.assertEqual(dec6(Decimal("1.2345678")), Decimal("1.234568"))

    def test_model_funcs(self):
        admin_log([self.user], "test msg 1")
        obj = LogEntry.objects.all().order_by("-pk").last()
        assert isinstance(obj, LogEntry)
        self.assertFalse(is_model_field_changed(obj, "change_message"))
        obj.change_message = "hello world"
        self.assertTrue(is_model_field_changed(obj, "change_message"))
        obj.save()
        self.assertFalse(is_model_field_changed(obj, "change_message"))
        obj2 = clone_model(obj)
        assert isinstance(obj2, LogEntry)
        self.assertEqual(obj.change_message, obj2.change_message)
        self.assertGreater(obj2.pk, obj.pk)
        label, val = get_model_field_label_and_value(obj, "action_time")
        self.assertEqual(label, _("action time"))
        self.assertEqual(str(obj.action_time), val)
        obj_b = get_object_or_none(obj.__class__, id=obj.id)
        self.assertEqual(obj_b.id, obj.id)
        obj_b = get_object_or_none(obj.__class__, id=-1)
        self.assertIsNone(obj_b)

    def test_format_table(self):
        a = [
            ["date", "description", "count", "unit price", "total price"],
            [date(2019, 12, 15), "oranges", 1000, dec2("0.99"), dec2("990.00")],
            [date(2020, 1, 3), "apples", 4, dec2("1.10"), dec2("4.40")],
            [date(2020, 11, 3), "apples", 5, dec2("10.10"), dec2("50.50")],
        ]
        out = format_table(a, has_label_row=True, max_col=10)
        out_ref = """
---------------------------------------------------
|      date|descript..|count|unit price|total pr..|
---------------------------------------------------
|2019-12-15|   oranges| 1000|      0.99|    990.00|
|2020-01-03|    apples|    4|      1.10|      4.40|
|2020-11-03|    apples|    5|     10.10|     50.50|
---------------------------------------------------
            """.strip()
        self.assertEqual(out, out_ref)

        out = format_table(a, has_label_row=True, max_col=20)
        out_ref = """
-----------------------------------------------------
|      date|description|count|unit price|total price|
-----------------------------------------------------
|2019-12-15|    oranges| 1000|      0.99|     990.00|
|2020-01-03|     apples|    4|      1.10|       4.40|
|2020-11-03|     apples|    5|     10.10|      50.50|
-----------------------------------------------------
            """.strip()
        self.assertEqual(out, out_ref)

        out = format_table(a, has_label_row=True, max_col=20, col_sep=" | ")
        out_ref = """
-------------------------------------------------------------
|      date | description | count | unit price | total price|
-------------------------------------------------------------
|2019-12-15 |     oranges |  1000 |       0.99 |      990.00|
|2020-01-03 |      apples |     4 |       1.10 |        4.40|
|2020-11-03 |      apples |     5 |      10.10 |       50.50|
-------------------------------------------------------------
            """.strip()
        self.assertEqual(out, out_ref)

        out = format_table(a, has_label_row=True, max_col=20, col_sep=" | ", left_align=[1])
        out_ref = """
-------------------------------------------------------------
|      date | description | count | unit price | total price|
-------------------------------------------------------------
|2019-12-15 | oranges     |  1000 |       0.99 |      990.00|
|2020-01-03 | apples      |     4 |       1.10 |        4.40|
|2020-11-03 | apples      |     5 |      10.10 |       50.50|
-------------------------------------------------------------
            """.strip()
        self.assertEqual(out, out_ref)

        out = format_table(a, has_label_row=True, max_col=20, col_sep=" | ", left_align=[1], max_line=50)
        out_ref = """
-------------------------------------------------
|      date | description | count | unit price|..
-------------------------------------------------
|2019-12-15 | oranges     |  1000 |       0.99|..
|2020-01-03 | apples      |     4 |       1.10|..
|2020-11-03 | apples      |     5 |      10.10|..
-------------------------------------------------
            """.strip()
        self.assertEqual(out, out_ref)

        out = format_table(a, left_align=[1], center_align=[0, 2, 3, 4], max_col=50)
        out_ref = """
-----------------------------------------------------
|   date   |description|count|unit price|total price|
|2019-12-15|oranges    |1000 |   0.99   |  990.00   |
|2020-01-03|apples     |  4  |   1.10   |   4.40    |
|2020-11-03|apples     |  5  |  10.10   |   50.50   |
-----------------------------------------------------
            """.strip()
        self.assertEqual(out, out_ref)

    def test_upper_lazy(self):
        s = gettext_lazy(upper_lazy("missing value"))
        s_ref = gettext_lazy("MISSING VALUE")
        s_en = "MISSING VALUE"
        s_fi = "PUUTTUVA ARVO"
        with override("fi"):
            self.assertEqual(s, s_fi)
            self.assertEqual(s, s_ref)
        with override("en"):
            self.assertEqual(s, s_en)
            self.assertEqual(s, s_ref)

    def test_capfirst_lazy(self):
        s = gettext_lazy(capfirst_lazy("missing value"))
        s_ref = gettext_lazy("Missing value")
        s_en = "Missing value"
        s_fi = "Puuttuva arvo"
        with override("fi"):
            self.assertEqual(s, s_fi)
            self.assertEqual(s, s_ref)
        with override("en"):
            self.assertEqual(s, s_en)
            self.assertEqual(s, s_ref)

    def test_media_paths(self):
        media_root1 = os.path.join(settings.MEDIA_ROOT, "path1/path2")
        media_root2 = os.path.join(settings.MEDIA_ROOT, "path3") + "/"
        test_paths = [
            (True, os.path.join(media_root1, "test1.file"), "path1/path2/test1.file"),
            (False, os.path.join("/diff/path", "test2.file"), "/diff/path/test2.file"),
            (True, os.path.join(media_root2, "test3.file"), "path3/test3.file"),
        ]
        for is_media, src, dst in test_paths:
            self.assertEqual(is_media_full_path(src), is_media)
            if is_media:
                self.assertEqual(strip_media_root(src), dst)
                self.assertEqual(get_media_full_path(dst), src)
                self.assertEqual(get_media_full_path(dst), get_media_full_path(get_media_full_path(dst)))

    def test_end_of_month(self):
        helsinki = pytz.timezone("Europe/Helsinki")
        # 1
        time_now = datetime(2020, 6, 5, 15, 47, 23, 818646)
        eom = end_of_month(time_now, tz=helsinki)
        eom_ref = helsinki.localize(datetime(2020, 6, 30, 23, 59, 59, 999999))
        self.assertEqual(eom, eom_ref)
        # 2
        time_now = datetime(2020, 7, 5, 15, 47, 23, 818646)
        eom = end_of_month(time_now, tz=helsinki)
        eom_ref = helsinki.localize(datetime(2020, 7, 31, 23, 59, 59, 999999))
        self.assertEqual(eom, eom_ref)
        # 3
        time_now = datetime(2020, 6, 5, 15, 47, 23, 818646)
        eom = end_of_month(time_now, n=1, tz=helsinki)
        eom_ref = helsinki.localize(datetime(2020, 7, 31, 23, 59, 59, 999999))
        self.assertEqual(eom, eom_ref)
        # 4
        time_now = datetime(2020, 7, 5, 15, 47, 23, 818646)
        eom = end_of_month(time_now, n=-2, tz=helsinki)
        eom_ref = helsinki.localize(datetime(2020, 5, 31, 23, 59, 59, 999999))
        self.assertEqual(eom, eom_ref)

    def test_iban_generator_and_validator(self):
        test_ibans = [
            "MD7289912714638112731113",
            "IS363252851674877586492113",
            "HR5125000099152386224",
            "CZ4750515755735423825528",
            "FI3253811381259333",
            "FR3212739000501869481882E94",
        ]
        for iban in test_ibans:
            iban_validator(iban)
        for cc in ["", "FI", "SE"]:
            for n in range(100):
                acc = iban_generator(cc)
                try:
                    iban_validator(acc)
                except Exception as e:
                    print("iban_generator() returned", acc, "but iban_validator() raised exception", e)
                    self.fail("iban_validator(iban_generator()) should not raise Exception, account number was {}".format(acc))
        with self.assertRaisesMessage(ValidationError, _("Invalid country code")):
            iban_generator("XX")
        with self.assertRaisesMessage(ValidationError, _("IBAN checksum generation does not support >26 character IBANs")):
            iban_generator("AL")

    def test_make_email_recipient(self):
        email_tests = [
            {
                "list": [
                    ("Jani Kajala", "kajala@example.com"),
                    '"Jani Kajala" <kajala@example.com>',
                    "<kajala@example.com>",
                    "kajala@example.com",
                ],
                "result": [
                    ("Jani Kajala", "kajala@example.com"),
                    ("Jani Kajala", "kajala@example.com"),
                    ("kajala@example.com", "kajala@example.com"),
                    ("kajala@example.com", "kajala@example.com"),
                ],
            }
        ]
        for et in email_tests:
            res = make_email_recipient_list(et["list"])
            self.assertListEqual(res, et["result"])

    def test_choices(self):
        val = choices_label(MY_CHOICES, MY_CHOICE_1)
        self.assertEqual(val, "MY_CHOICE_1")

    def test_camel_case(self):
        pairs = [
            ("camelCaseWord", "camel_case_word"),
            ("camelCase", "camel_case"),
            ("camel", "camel"),
            ("camelCCase", "camel_c_case"),
        ]
        for cc, us in pairs:
            self.assertEqual(camel_case_to_underscore(cc), us)
            self.assertEqual(cc, underscore_to_camel_case(us))

    def create_dummy_request(self, path: str = "/admin/login/"):
        request = request_factory.get(path)
        request.user = self.user  # type: ignore
        request.user.profile = DummyUserProfile()  # type: ignore
        return request

    def test_model_admin_base(self):
        # test that actions sorting by name works
        request = self.create_dummy_request()
        user = self.user
        model_admin = MyCustomAdmin(LogEntry, admin.site)
        res = model_admin.get_actions(request)
        self.assertEqual(list(res.items())[0][0], "dummy_admin_func_a", "ModelAdminBase.get_actions sorting failed")
        self.assertEqual(list(res.items())[1][0], "dummy_admin_func_b", "ModelAdminBase.get_actions sorting failed")

        # create 10 LogEntry for test user, 5 with text "VisibleLogMessage" and 5 "InvisibleLogMessage"
        # then check that "VisibleLogMessage" log entries are not visible since max_history_length = 5
        LogEntry.objects.filter(object_id=user.id).delete()
        for n in range(5):
            admin_log([user], "InvisibleLogMessage")
        for n in range(5):
            admin_log([user], "VisibleLogMessage")
        self.assertEqual(LogEntry.objects.filter(object_id=user.id).count(), 10)
        history_url = "/admin/auth/user/{}/history/".format(user.id)
        c = self.client
        c.get(history_url, follow=True)
        c.post("/admin/login/", {"username": "test@example.com", "password": "test1234"})
        res = c.get(history_url)
        content = res.content.decode()
        assert isinstance(content, str)
        self.assertEqual(content.count("VisibleLogMessage"), 5)
        self.assertEqual(content.count("InvisibleLogMessage"), 0)

    def test_auth(self):
        req = self.create_dummy_request()
        get_auth_user(req)  # type: ignore
        req.user = None
        self.assertIsNone(get_auth_user_or_none(req))
        try:
            get_auth_user(req)  # type: ignore
            self.fail("get_auth_user fail")
        except NotAuthenticated:
            pass
        try:
            model_admin = AuthUserMixin()
            model_admin.request = req
            user = model_admin.auth_user  # noqa
            self.fail("get_auth_user fail")
        except NotAuthenticated:
            pass

    def test_middleware(self):
        # EnsureOriginMiddleware
        request = self.create_dummy_request("/admin/login/")
        EnsureOriginMiddleware(dummy_middleware_get_response)(request)
        self.assertIn("HTTP_ORIGIN", request.META)
        self.assertEqual(request.META["HTTP_ORIGIN"], request.get_host())

        # LogExceptionMiddleware
        dummy_log = DummyLogHandler()
        jutil_middleware_logger.addHandler(dummy_log)
        try:
            raise Exception("Dummy exception, ignore this")  # noqa
        except Exception as e:
            mw = LogExceptionMiddleware(dummy_middleware_get_response)
            mw.process_exception(request, e)
        msg = dummy_log.msgs.pop()
        self.assertIn("Exception: Dummy exception, ignore this", msg)
        self.assertIn("user=test@example.com", msg)
        jutil_middleware_logger.removeHandler(dummy_log)

        # EnsureLanguageCookieMiddleware
        lang_cookie_tests = [
            ("/admin/login/", settings.LANGUAGE_CODE),
            ("/admin/login/?django_language=fi", "fi"),
        ]
        for request_path, lang_code in lang_cookie_tests:
            request = self.create_dummy_request(request_path)
            mw = EnsureLanguageCookieMiddleware(dummy_middleware_get_response)
            res = mw(request)
            assert isinstance(res, HttpResponse)
            self.assertEqual(res.status_code, 200)
            self.assertIn("django_language", request.COOKIES)
            self.assertIn(request.COOKIES["django_language"], lang_code)
            self.assertIn("django_language", res.cookies)
            self.assertEqual(str(res.cookies["django_language"]), "Set-Cookie: django_language={}; Path=/".format(lang_code))

        # ActivateUserProfileTimezoneMiddleware
        user = self.user
        self.assertTrue(user.is_authenticated)
        user.profile.timezone = "Europe/Helsinki"
        with timezone.override(pytz.timezone("America/Chicago")):
            request = self.create_dummy_request("/admin/login/")
            mw = ActivateUserProfileTimezoneMiddleware(dummy_time_zone_response)
            res = mw(request)
            content = res.content.decode()
            self.assertEqual(content, user.profile.timezone)

    def test_safe_fields(self):
        class TestModel(models.Model):
            cf = SafeCharField(max_length=256)
            tf = SafeTextField()

        obj = TestModel()
        data = 'hello world <script>alert("popup")<script>'
        data_ref = "hello world "
        for f in obj._meta.fields:
            if f.name in ["cf", "tf"]:
                f.save_form_data(obj, data)
        self.assertEqual(obj.cf, data_ref)
        self.assertEqual(obj.tf, data_ref)

    def test_responses(self):
        a = [
            ["date", "description", "count", "unit price", "total price"],
            [date(2019, 12, 15), "oranges", 1000, dec2("0.99"), dec2("990.00")],
            [date(2020, 1, 3), "apples", 4, dec2("1.10"), dec2("4.40")],
            [date(2020, 11, 3), "apples", 5, dec2("10.10"), dec2("50.50")],
        ]
        res = CsvResponse(a, "test.csv")
        content_ref = b"date,description,count,unit price,total price\r\n2019-12-15,oranges,1000,0.99,990.00\r\n2020-01-03,apples,4,1.10,4.40\r\n2020-11-03,apples,5,10.10,50.50\r\n"  # noqa
        self.assertEqual(content_ref, res.content)
        print(res.content.decode())

    def test_wait_object_or_none(self):
        admin_log([self.user], "Hello, world")
        e = LogEntry.objects.all().filter(object_id=self.user.id).last()
        assert isinstance(e, LogEntry)
        e_id = e.id
        obj = wait_object_or_none(LogEntry, id=e_id)
        self.assertIsNotNone(obj)
        self.assertEqual(obj.id, e_id)
        t0 = now()
        obj = wait_object_or_none(LogEntry, timeout=1.0, sleep_interval=0.1, id=e_id + 1)
        t1 = now()
        self.assertIsNone(obj)
        self.assertGreater(t1 - t0, timedelta(seconds=0.99))

    def test_parse_bool(self):
        refs_true = [
            True,
            "True",
            "1",
            1,
            "true",
            "yes",
        ]
        refs_false = [
            False,
            "False",
            "0",
            0,
            "false",
            "no",
        ]
        for ref in refs_true:
            self.assertTrue(parse_bool(ref))
        for ref in refs_false:
            self.assertFalse(parse_bool(ref))

    def test_exception_convert(self):
        e = transform_exception_to_drf(ValidationError({"hello": "world"}))
        assert isinstance(e, DRFValidationError)
        self.assertTrue(isinstance(e, DRFValidationError))
        self.assertIn("hello", e.detail)  # type: ignore
        self.assertEqual(e.detail["hello"], [ErrorDetail(string="world", code="invalid")])  # type: ignore

    def test_list_files(self):
        dir_name = os.path.join(settings.BASE_DIR, "jutil/locale")
        cases = [
            [
                {"recurse": True, "use_media_root": True, "suffix": ".PO", "ignore_case": True},
                ["jutil/locale/fi/LC_MESSAGES/django.po", "jutil/locale/sv/LC_MESSAGES/django.po", "jutil/locale/en/LC_MESSAGES/django.po"],
            ],
            [
                {"recurse": True, "use_media_root": True, "suffix": ".po", "ignore_case": False},
                ["jutil/locale/fi/LC_MESSAGES/django.po", "jutil/locale/sv/LC_MESSAGES/django.po", "jutil/locale/en/LC_MESSAGES/django.po"],
            ],
            [
                {"recurse": True, "use_media_root": True, "suffix": ".PO", "ignore_case": False},
                [],
            ],
            [
                {"recurse": False, "use_media_root": True, "suffix": ".PO", "ignore_case": True},
                [],
            ],
        ]
        for call_kw, data_ref in cases:
            out = StringIO()
            call_command("list_files", dir_name, json=True, stdout=out, **call_kw)
            data = json.loads(out.getvalue())
            self.assertListEqual(data_ref, data)

    def test_filters(self):
        vals = [
            (
                "24: REGISTRO COMPLEMENTARIO DE INFORMACIÓN DE EQUIVALENCIA DE IMPORTE DEL APUNTE (OPCIONAL)",
                "_24_REGISTRO_COMPLEMENTARIO_DE_INFORMACION_DE_EQUIVALENCIA_DE_IMPORTE_DEL_APUNTE_OPCIONAL",
            ),
            (
                "a/12/312/2/12--12-/123",
                "a1231221212123",
            ),
        ]
        for src, dst in vals:
            self.assertEqual(variable_name_sanitizer(src), dst)

    def test_find_file(self):
        dir_name = os.path.join(settings.BASE_DIR, "jutil/locale")
        cases = [
            [
                {"recurse": True, "use_media_root": True, "filename": "django.po"},
                ["jutil/locale/fi/LC_MESSAGES/django.po", "jutil/locale/sv/LC_MESSAGES/django.po", "jutil/locale/en/LC_MESSAGES/django.po"],
            ],
            [
                {"recurse": False, "use_media_root": True, "filename": "django.po"},
                [],
            ],
            [
                {"recurse": True, "use_media_root": True, "filename": "en/LC_MESSAGES/django.po"},
                ["jutil/locale/en/LC_MESSAGES/django.po"],
            ],
        ]
        for call_kw, data_ref in cases:
            data = find_file(**call_kw, dir_name=dir_name)
            self.assertListEqual(data_ref, data)

    def test_command_utils(self):
        for cmd_name in ["apps", "geo_ip", "list_files", "send_email", "setpass"]:
            cls = get_command_by_name(cmd_name)
            self.assertTrue(isinstance(cls, BaseCommand))
            self.assertEqual(get_command_name(cls), cmd_name)

    def test_redis_helpers(self):
        for k in ["jani", "jani2", "jani3", "jani4"]:
            redis_delete(k)
        self.assertEqual(redis_get_json_or_none("jani2"), None)
        self.assertEqual(redis_get_bytes_or_none("jani"), None)
        redis_set_json("jani", {"a": 1})
        redis_set_json("jani2", "testing")
        redis_set_bytes("jani3", b'{"a":2}')
        self.assertDictEqual(redis_get_json("jani"), {"a": 1})
        self.assertEqual(redis_get_json("jani2"), "testing")
        self.assertEqual(redis_get_bytes("jani"), b'{"a": 1}')
        self.assertEqual(redis_get_bytes("jani2"), b'"testing"')
        self.assertDictEqual(redis_get_json("jani3"), {"a": 2})

    def test_openpyxl_helpers(self):
        from jutil.openpyxl_helpers import rows_to_workbook, save_workbook_to_bytes  # type: ignore  # noqa
        from openpyxl import load_workbook  # type: ignore  # noqa

        rows = [["hello", "world"], [True, 1.0, 2]]
        wb = rows_to_workbook(rows)
        buf = save_workbook_to_bytes(wb)
        wb2 = load_workbook(filename=BytesIO(buf))
        self.assertEqual(wb2.active.cell(1, 1).value, "hello")
        self.assertEqual(wb2.active.cell(2, 2).value, 1.0)
        self.assertEqual(wb2.active.cell(2, 3).value, 2)

    def test_is_iban(self):
        valid_ibans = ["AL35202111090000000001234567", "AD1400080001001234567890", "AT483200000012345864"]
        for valid_iban in valid_ibans:
            self.assertTrue(is_iban(valid_iban))
        invalid_ibans = ["GB96BARC202015300934591", "US64SVBKUS6S3300958879"]
        for invalid_iban in invalid_ibans:
            self.assertFalse(is_iban(invalid_iban))

    def test_alnum_filter(self):
        results = [
            ("!@3123a bcd222", "", "3123abcd222"),
            ("hello world", " ", "hello world"),
        ]
        for input_value, extra, correct_result in results:
            self.assertEqual(alnum_filter(input_value, extra=extra), correct_result)


dummy_admin_func_a.short_description = "A"  # type: ignore
dummy_admin_func_b.short_description = "B"  # type: ignore

admin.site.unregister(User)
admin.site.register(User, MyCustomAdmin)
admin.site.register(LogEntry, MyCustomAdmin)
