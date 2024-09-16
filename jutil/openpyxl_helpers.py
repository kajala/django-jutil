from datetime import datetime, date
from decimal import Decimal
from tempfile import NamedTemporaryFile
from zipfile import ZipFile, ZIP_DEFLATED
from django.http import HttpResponse
from django.utils.html import strip_tags
from typing import Any, List, Optional, Dict
import os

from openpyxl.utils import column_index_from_string

try:
    import openpyxl  # type: ignore

    if int(openpyxl.__version__.split(".", maxsplit=1)[0]) < 3:
        raise Exception("Invalid version")  # noqa
except Exception as err:
    raise Exception("Using jutil.openpyxl_helpers requires openpyxl>3.0 installed") from err  # noqa

from openpyxl import Workbook  # type: ignore
from openpyxl.styles import Alignment, NamedStyle  # type: ignore
from openpyxl.worksheet.worksheet import Worksheet  # type: ignore
from openpyxl.writer.excel import ExcelWriter  # type: ignore


class CellConfig:
    decimal_format: str = "0.00"
    int_format: str = "0"
    float_format: str = "0.0"
    number_alignment = Alignment(horizontal="right")
    date_style = NamedStyle(name="datetime", number_format="YYYY-MM-DD")


class ExcelResponse(HttpResponse):
    def __init__(self, book: Workbook, filename: str, disposition: str = "", content_type: str = ""):
        if not disposition:
            disposition = "filename=" + os.path.basename(filename)
        if not content_type:
            content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        super().__init__(
            content=save_workbook_to_bytes(book),
            content_type=content_type,
        )
        self["Content-Disposition"] = disposition


def save_workbook_to_bytes(workbook: Workbook) -> bytes:
    """Return an in-memory workbook."""
    with NamedTemporaryFile() as tmp:
        with ZipFile(tmp, "w", ZIP_DEFLATED, allowZip64=True) as archive:
            writer = ExcelWriter(workbook, archive)
            writer.save()
            tmp.seek(0)
            virtual_workbook = tmp.read()
            archive.close()
        tmp.close()
    return virtual_workbook


def set_cell_value(sheet: Worksheet, row_index: int, column_index: int, val: Any, config: Optional[CellConfig] = None):
    if config is None:
        config = CellConfig()
    c = sheet.cell(row_index + 1, column_index + 1)
    if isinstance(val, Decimal):
        c.number_format = config.decimal_format
        c.alignment = config.number_alignment
    elif isinstance(val, int):
        c.number_format = config.int_format
        c.alignment = config.number_alignment
    elif isinstance(val, float):
        c.number_format = config.float_format
        c.alignment = config.number_alignment
    elif isinstance(val, (datetime, date)):
        c.style = config.date_style
        if isinstance(val, datetime):
            val = val.replace(tzinfo=None)
    else:
        if str(val).startswith("https://"):
            c.hyperlink = str(val)  # type: ignore
        val = strip_tags(str(val if val else ""))
    c.value = val
    return c


def get_sheet_column_letters(sheet: Worksheet) -> List[str]:
    column_letters = list(openpyxl.utils.get_column_letter(col_number + 1) for col_number in range(sheet.max_column))
    return column_letters


def rows_to_sheet(sheet: Worksheet, rows: List[List[Any]], config: Optional[CellConfig]):
    if config is None:
        config = CellConfig()
    for row_ix, row in enumerate(list(rows)):
        for col_ix, val in enumerate(list(row)):
            set_cell_value(sheet, row_ix, col_ix, val, config)


def rows_to_workbook(rows: List[List[Any]], config: Optional[CellConfig] = None) -> Workbook:
    if config is None:
        config = CellConfig()
    book = Workbook()
    sheet = book.active
    assert isinstance(sheet, Worksheet)
    rows_to_sheet(sheet, rows, config)
    return book


def set_sheet_column_widths(sheet: Worksheet, column_letter_width_pairs: Dict[str, int]):
    """
    Sets worksheet column widths.

    For example:
    set_sheet_column_widths(wb.active, {"A": 20, "B": 20, "C": 20})

    Args:
        sheet: Worksheet
        column_letter_width_pairs: Column letter - width (int) dict

    Returns:

    """
    for column_letter, column_width in column_letter_width_pairs.items():
        sheet.column_dimensions[column_letter].width = column_width


def set_sheet_column_alignments(sheet: Worksheet, column_letter_alignment_pairs: Dict[str, str]):
    """
    Sets worksheet column alignments.
    Alignment value must be one of ‘fill’, ‘left’, ‘distributed’, ‘justify’, ‘center’, ‘general’, ‘centerContinuous’, ‘right’.

    For example:
    set_sheet_column_alignments(wb.active, {"A": "left", "B": "right", "C": "right"})

    Args:
        sheet: Worksheet
        column_letter_alignment_pairs: Column letter - alignment dict.

    Returns:

    """
    for row in sheet:
        for column_letter, alignment in column_letter_alignment_pairs.items():
            column_ix = column_index_from_string(column_letter)
            row[column_ix - 1].alignment = Alignment(horizontal=alignment)


def find_sheet_column_index_by_label(sheet: Worksheet, label: str, label_row_index: int = 0) -> Optional[int]:
    row_num = label_row_index + 1
    for col_num in range(1, sheet.max_column + 1):
        cell = sheet.cell(row_num, col_num)
        if cell.value is None:
            break
        if str(cell.value) == label:
            return col_num - 1
    return None
